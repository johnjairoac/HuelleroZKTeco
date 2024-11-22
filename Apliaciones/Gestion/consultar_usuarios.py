import tkinter as tk
from tkinter import ttk
import sqlite3
import pandas as pd
from DB.db_config import DB_PATH
from tkinter import messagebox
from datetime import datetime
import os
from DescargasExcel.excel_config import EXCEL_PATH  # Importar la ruta de guardado desde el archivo de configuración
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Función para conectar a la base de datos SQLite
def conectar_db():
    conn = sqlite3.connect(DB_PATH)  # Usamos la ruta de la base de datos definida en DB_PATH
    return conn

# Función para obtener todos los usuarios de la base de datos
def obtener_todos_los_usuarios(filtro=""):
    conn = conectar_db()
    cursor = conn.cursor()
    if filtro:
        cursor.execute('''SELECT id_empleado, nombre, apellido, cedula FROM empleados WHERE nombre LIKE ? OR id_empleado LIKE ?''', ('%' + filtro + '%', '%' + filtro + '%'))
    else:
        cursor.execute('''SELECT id_empleado, nombre, apellido, cedula FROM empleados''')
    usuarios = cursor.fetchall()
    conn.close()
    return usuarios

# Función para actualizar la vista de los usuarios en el Treeview
def actualizar_usuarios(treeview, filtro=""):
    # Eliminar todas las filas existentes
    for row in treeview.get_children():
        treeview.delete(row)
    # Obtener la lista de usuarios filtrada
    usuarios = obtener_todos_los_usuarios(filtro)
    # Insertar los nuevos usuarios en el Treeview
    for usuario in usuarios:
        treeview.insert("", tk.END, values=usuario)

# Función para exportar usuarios a un archivo Excel
def exportar_a_excel():
    # Obtener todos los usuarios
    usuarios = obtener_todos_los_usuarios()

    # Crear un DataFrame de pandas con los datos de los usuarios
    df = pd.DataFrame(usuarios, columns=["ID Empleado", "Nombre", "Apellido", "Cédula"])

    # Generar el nombre de archivo con fecha y hora actual
    fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    archivo_excel = f"usuarios_registrados_{fecha_hora}.xlsx"

    # Crear la ruta completa para el archivo Excel
    ruta_completa = os.path.join(EXCEL_PATH, archivo_excel)

    # Verificar si la carpeta de destino existe, si no, crearla
    if not os.path.exists(EXCEL_PATH):
        os.makedirs(EXCEL_PATH)

    # Guardar el DataFrame en un archivo Excel
    try:
        df.to_excel(ruta_completa, index=False)

        # Abrir el archivo Excel con openpyxl para ajustarlo
        wb = load_workbook(ruta_completa)
        sheet = wb.active

        # Ajustar el contenido de las celdas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Aplicar color verde a las celdas de la fila 1, columnas de la A a la D
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        for cell in sheet["1:1"]:
            cell.fill = green_fill

        # Guardar los cambios realizados
        wb.save(ruta_completa)

        # Abrir automáticamente el archivo Excel
        os.startfile(ruta_completa)

        messagebox.showinfo("Éxito", f"El archivo Excel ha sido generado y abierto correctamente: {ruta_completa}")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al generar el archivo Excel: {e}")

# Ventana para consultar todos los usuarios
def ventana_consultar_usuarios():
    ventana = tk.Tk()
    ventana.title("Consultar Usuarios")
    ventana.configure(bg="#2C2F38")  # Fondo oscuro

    # Obtener el tamaño de la pantalla
    screen_width = ventana.winfo_screenwidth()
    screen_height = ventana.winfo_screenheight()

    # Establecer el tamaño de la ventana
    window_width = 1000
    window_height = 600

    # Calcular la posición de la ventana para centrarla
    position_top = int(screen_height / 2 - window_height / 2)
    position_left = int(screen_width / 2 - window_width / 2)

    # Establecer la geometría de la ventana (centrada en la pantalla)
    ventana.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")

    # Crear un contenedor (Frame) para los elementos
    contenedor = tk.Frame(ventana, bg="#2C2F38")
    contenedor.pack(padx=20, pady=20, fill="both", expand=True)

    # Etiqueta principal
    tk.Label(contenedor, text="Usuarios Registrados", fg="white", bg="#2C2F38", font=("Arial", 14, "bold")).pack(pady=10)

    # Campo de entrada para el filtro (nombre o ID)
    filtro_label = tk.Label(contenedor, text="Filtrar por Nombre o ID:", fg="white", bg="#2C2F38", font=("Arial", 12))
    filtro_label.pack(pady=5)

    filtro_entry = tk.Entry(contenedor, font=("Arial", 12), bg="#34495E", fg="white", width=50)
    filtro_entry.pack(pady=5)

    # Función para filtrar cuando se presiona Enter
    def aplicar_filtro(event=None):
        filtro = filtro_entry.get()
        actualizar_usuarios(treeview, filtro)

    filtro_entry.bind('<Return>', aplicar_filtro)

    # Crear un marco para el Treeview y su scrollbar
    frame_treeview = tk.Frame(contenedor, bg="#2C2F38")
    frame_treeview.pack(pady=10, fill="both", expand=True)

    # Crear un Treeview para mostrar los usuarios en formato tabla
    columnas = ("ID Empleado", "Nombre", "Apellido", "Cédula")
    treeview = ttk.Treeview(frame_treeview, columns=columnas, show="headings", height=15)
    treeview.pack(side="left", fill="both", expand=True)

    # Configurar las columnas del Treeview
    for col in columnas:
        treeview.heading(col, text=col)
        treeview.column(col, width=200, anchor="center")

    # Barra de desplazamiento vertical dentro del marco del Treeview
    scrollbar = ttk.Scrollbar(frame_treeview, orient="vertical", command=treeview.yview)
    treeview.config(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    # Marco para los botones alineados horizontalmente
    frame_botones = tk.Frame(contenedor, bg="#2C2F38")
    frame_botones.pack(pady=20)

    # Botón de actualizar lista
    boton_actualizar = tk.Button(frame_botones, text="Actualizar Lista", command=lambda: actualizar_usuarios(treeview), font=("Arial", 14, "bold"),
                                 bg="#3498DB", fg="white", relief="flat", activebackground="#2980B9", width=20, height=2)
    boton_actualizar.pack(side="left", padx=10)

    # Botón de exportar a Excel
    boton_excel = tk.Button(frame_botones, text="📊 Exportar a Excel", command=exportar_a_excel, font=("Arial", 14, "bold"),
                            bg="#28B463", fg="white", relief="flat", activebackground="#1D8348", width=20, height=2)
    boton_excel.pack(side="left", padx=10)

    # Llamada inicial para cargar los usuarios
    actualizar_usuarios(treeview)

    ventana.mainloop()

# Llamada para abrir la ventana solo si se ejecuta directamente
if __name__ == "__main__":
    ventana_consultar_usuarios()
