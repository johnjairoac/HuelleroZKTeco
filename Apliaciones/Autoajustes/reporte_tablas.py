import sqlite3
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime
import os
from tkcalendar import DateEntry
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from DB.db_config import DB_PATH

# Diccionario para traducir los días de la semana al español
dias_semana_es = {
    'Monday': 'Lunes',
    'Tuesday': 'Martes',
    'Wednesday': 'Miércoles',
    'Thursday': 'Jueves',
    'Friday': 'Viernes',
    'Saturday': 'Sábado',
    'Sunday': 'Domingo'
}

# Función para conectar a la base de datos SQLite
def conectar_db():
    conn = sqlite3.connect(DB_PATH)
    return conn

# Función para verificar si una fecha es festiva
def es_festivo(fecha):
    conn = conectar_db()
    cursor = conn.cursor()
    consulta = "SELECT 1 FROM festivos WHERE fecha = ?"
    cursor.execute(consulta, (fecha,))
    resultado = cursor.fetchone()
    conn.close()
    return "Sí" if resultado else "No"

# Función para obtener datos de las tablas empleados y archivos_extraidos
def obtener_empleados_archivos(fecha_inicio=None, fecha_fin=None, id_empleado=None):
    conn = conectar_db()
    cursor = conn.cursor()

    consulta = '''
        SELECT e.id_empleado, e.nombre, e.apellido, e.cedula, a.date, a.time
        FROM empleados e
        JOIN archivos_extraidos a ON e.cedula = a.user_name
    '''

    condiciones = []
    if fecha_inicio and fecha_fin:
        condiciones.append(f"a.date BETWEEN '{fecha_inicio}' AND '{fecha_fin}'")
    if id_empleado:
        condiciones.append(f"e.id_empleado = {id_empleado}")

    if condiciones:
        consulta += " WHERE " + " AND ".join(condiciones)

    cursor.execute(consulta)
    datos = cursor.fetchall()
    conn.close()
    return datos

# Función para exportar datos a un archivo Excel
def exportar_a_excel_empleados_archivos(fecha_inicio=None, fecha_fin=None, id_empleado=None):
    datos = obtener_empleados_archivos(fecha_inicio, fecha_fin, id_empleado)
    columnas = ["ID Empleado", "Nombre", "Apellido", "Cédula", "Fecha", "Hora", "Día de la Semana", "Festivo"]

    datos_con_dia = []
    for fila in datos:
        fecha = fila[4]
        fecha_objeto = datetime.strptime(fecha, '%Y-%m-%d')
        dia_semana = dias_semana_es.get(fecha_objeto.strftime('%A'), fecha_objeto.strftime('%A'))
        festivo = es_festivo(fecha)
        datos_con_dia.append(fila + (dia_semana, festivo))

    df = pd.DataFrame(datos_con_dia, columns=columnas)
    fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    archivo_excel = f"empleados_archivos_{fecha_hora}.xlsx"

    try:
        df.to_excel(archivo_excel, index=False, sheet_name="Empleados", engine='openpyxl')
        wb = load_workbook(archivo_excel)
        ws = wb.active

        # Aplicar estilo al encabezado
        amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col in range(1, len(columnas) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = amarillo
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Autoajustar columnas y centrar contenido
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # Obtener letra de la columna
            for cell in col:
                try:
                    if cell.value:  # Obtener el valor de la celda
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Ajustar un poco el ancho
            ws.column_dimensions[col_letter].width = adjusted_width
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(archivo_excel)
        messagebox.showinfo("Éxito", f"El archivo Excel ha sido generado correctamente: {archivo_excel}")
        os.startfile(archivo_excel)
    except Exception as e:
        messagebox.showerror("Error", f"Error al generar el archivo Excel: {e}")






# Función para mostrar los datos en el Treeview
def mostrar_datos(treeview, fecha_inicio, fecha_fin, id_empleado):
    for row in treeview.get_children():
        treeview.delete(row)

    datos = obtener_empleados_archivos(fecha_inicio, fecha_fin, id_empleado)
    for dato in datos:
        fecha = dato[4]
        fecha_objeto = datetime.strptime(fecha, '%Y-%m-%d')
        dia_semana = dias_semana_es.get(fecha_objeto.strftime('%A'), fecha_objeto.strftime('%A'))
        festivo = es_festivo(fecha)
        treeview.insert("", tk.END, values=dato + (dia_semana, festivo))

# Función para centrar la ventana
def centrar_ventana(ventana, width, height):
    screen_width = ventana.winfo_screenwidth()
    screen_height = ventana.winfo_screenheight()
    position_top = int(screen_height / 2 - height / 2)
    position_right = int(screen_width / 2 - width / 2)
    ventana.geometry(f'{width}x{height}+{position_right}+{position_top}')

# Interfaz gráfica
def ventana_consultar_empleados_archivos():
    ventana = tk.Tk()
    ventana.title("Consultar Archivos Extraidos")
    ventana.configure(bg="#2C2F38")

    centrar_ventana(ventana, 1000, 800)

    tk.Label(ventana, text="Personal y Documentos Procesados", fg="white", bg="#2C2F38", font=("Arial", 14, "bold")).pack(pady=10)

    frame = tk.Frame(ventana)
    frame.pack(pady=10, fill="both", expand=True)

    frame_fechas = tk.Frame(ventana)
    frame_fechas.pack(pady=10)

    contenedor_fechas = tk.Frame(frame_fechas, bg="#2C2F38")
    contenedor_fechas.pack()

    tk.Label(contenedor_fechas, text="Fecha de inicio:", fg="white", bg="#2C2F38").pack(side="left", padx=5)
    fecha_inicio_entry = DateEntry(contenedor_fechas, width=12, background="darkblue", foreground="white", date_pattern="yyyy-mm-dd")
    fecha_inicio_entry.pack(side="left", padx=10)

    tk.Label(contenedor_fechas, text="Fecha de fin:", fg="white", bg="#2C2F38").pack(side="left", padx=5)
    fecha_fin_entry = DateEntry(contenedor_fechas, width=12, background="darkblue", foreground="white", date_pattern="yyyy-mm-dd")
    fecha_fin_entry.pack(side="left", padx=10)

    tk.Label(ventana, text="ID de Empleado (opcional):", fg="white", bg="#2C2F38").pack(pady=10)
    id_empleado_entry = tk.Entry(ventana, font=("Arial", 12), width=20)
    id_empleado_entry.pack(pady=5)

    columnas = ["ID Empleado", "Nombre", "Apellido", "Cédula", "Fecha", "Hora", "Día de la Semana", "Festivo"]

    tree_frame = tk.Frame(frame)
    tree_frame.pack(fill="both", expand=True)

    scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
    scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
    treeview = ttk.Treeview(tree_frame, columns=columnas, show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    scroll_y.config(command=treeview.yview)
    scroll_x.config(command=treeview.xview)

    scroll_y.pack(side="right", fill="y")
    scroll_x.pack(side="bottom", fill="x")
    treeview.pack(fill="both", expand=True)

    for col in columnas:
        treeview.heading(col, text=col)
        treeview.column(col, anchor=tk.CENTER, width=120)

    def actualizar_datos():
        fecha_inicio = fecha_inicio_entry.get()
        fecha_fin = fecha_fin_entry.get()
        id_empleado = id_empleado_entry.get()
        mostrar_datos(treeview, fecha_inicio, fecha_fin, id_empleado if id_empleado else None)

    tk.Button(ventana, text="Consultar", command=actualizar_datos, bg="#4CAF50", fg="white", font=("Arial", 12, "bold")).pack(pady=10)

    def exportar_excel():
        fecha_inicio = fecha_inicio_entry.get()
        fecha_fin = fecha_fin_entry.get()
        id_empleado = id_empleado_entry.get()
        exportar_a_excel_empleados_archivos(fecha_inicio, fecha_fin, id_empleado if id_empleado else None)

    tk.Button(ventana, text="Exportar a Excel", command=exportar_excel, bg="#007BFF", fg="white", font=("Arial", 12, "bold")).pack(pady=10)

    ventana.mainloop()

if __name__ == "__main__":
    ventana_consultar_empleados_archivos()
