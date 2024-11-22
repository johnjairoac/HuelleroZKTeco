import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime, timedelta

# Función para conectar a la base de datos SQLite
from DB.db_config import DB_PATH
import sqlite3


def conectar_db():
    conn = sqlite3.connect(DB_PATH)
    return conn


# Función para calcular el tiempo total trabajado en horas (ignorando los segundos)
def calcular_tiempo_total(hora_inicio, hora_fin):
    formato_hora = "%H:%M"  # Asumimos que la hora está en formato HH:MM
    try:
        # Eliminar los segundos (si los hay) y procesar solo horas y minutos
        hora_inicio = hora_inicio.split(":")[:2]  # Solo horas y minutos
        hora_fin = hora_fin.split(":")[:2]  # Solo horas y minutos

        # Convertir las horas a objetos datetime, ignorando los segundos
        inicio = datetime.strptime(":".join(hora_inicio), formato_hora)
        fin = datetime.strptime(":".join(hora_fin), formato_hora)

        # Si la hora de fin es anterior a la de inicio, significa que pasó a otro día (turno nocturno)
        if fin < inicio:
            fin += pd.Timedelta(days=1)

        # Calcular la diferencia en horas
        tiempo_total = (fin - inicio).seconds / 3600  # Convertir segundos a horas
        return round(tiempo_total, 2)  # Redondeamos a 2 decimales para mayor precisión
    except Exception as e:
        print(f"Error al calcular el tiempo: {e}")
        return 0  # En caso de error, retornamos 0 como valor de tiempo


# Función para convertir el tiempo en formato HH:MM a minutos
def convertir_a_minutos(horas_y_minutos):
    try:
        horas, minutos = map(int, horas_y_minutos.split(':'))
        total_minutos = horas * 60 + minutos
        return total_minutos
    except ValueError:
        return 0  # En caso de error, retornamos 0


# Función para restar 30 minutos de un tiempo dado en formato HH:MM
def restar_30_minutos(tiempo, dia_semana):
    try:
        # Si es sábado o domingo, no restar 30 minutos
        if dia_semana in ['Sábado', 'Domingo']:
            return tiempo  # No restamos los 30 minutos

        # Convertir el tiempo a minutos
        tiempo_en_minutos = convertir_a_minutos(tiempo)
        # Restar 30 minutos
        tiempo_limpio_en_minutos = tiempo_en_minutos - 30
        # Convertir de nuevo a formato HH:MM
        horas = tiempo_limpio_en_minutos // 60
        minutos = tiempo_limpio_en_minutos % 60
        return f"{horas:02}:{minutos:02}"  # Devolver el tiempo con formato HH:MM
    except Exception as e:
        print(f"Error al restar 30 minutos: {e}")
        return "00:00"  # En caso de error, retornamos 00:00


# Función para consultar los días festivos desde la base de datos
def obtener_festivos_db():
    try:
        conn = conectar_db()  # Conectar a la base de datos
        cursor = conn.cursor()
        cursor.execute("SELECT fecha FROM festivos")  # Obtener todas las fechas festivas
        festivos = cursor.fetchall()  # Obtener todas las filas
        conn.close()

        # Convertir la lista de fechas festivas a un conjunto de fechas
        festivos_set = {str(festivo[0]) for festivo in festivos}  # Convertir a formato 'YYYY-MM-DD'
        return festivos_set
    except Exception as e:
        print(f"Error al obtener festivos: {e}")
        return set()  # En caso de error, retornamos un conjunto vacío


# Función para agregar la columna "Festivos"
def agregar_columna_festivos(df):
    festivos_db = obtener_festivos_db()  # Obtener los festivos desde la base de datos

    # Crear la columna "Festivos" que indica si la fecha es festiva o no
    df['Festivos'] = df['Fecha'].apply(lambda x: 'Sí' if str(x) in festivos_db else 'No')
    return df

# Función para obtener los ciclos desde la base de datos
def obtener_ciclos_db():
    try:
        conn = conectar_db()  # Conectar a la base de datos
        cursor = conn.cursor()
        cursor.execute("SELECT tipo_ciclo, hora_inicio, hora_fin FROM ciclos")  # Obtener los ciclos
        ciclos = cursor.fetchall()  # Obtener todas las filas
        conn.close()

        # Crear una lista con los ciclos (hora_inicio, hora_fin, tipo_ciclo)
        ciclos_list = [(str(ciclo[1]), str(ciclo[2]), ciclo[0]) for ciclo in ciclos]
        return ciclos_list
    except Exception as e:
        print(f"Error al obtener ciclos: {e}")
        return []  # En caso de error, retornamos una lista vacía



# Función para determinar el ciclo de acuerdo a la hora
# Función para determinar el ciclo de acuerdo a la hora
def determinar_ciclo(hora, ciclos):
    try:
        # Truncar cualquier dato extra (como segundos) de la hora
        hora = hora.split(':')[:2]  # Tomar solo la parte de horas y minutos
        hora = ":".join(hora)  # Reunir de nuevo la hora en formato HH:MM

        # Convertir la hora a un objeto datetime
        hora_obj = datetime.strptime(hora, "%H:%M")

        for hora_inicio, hora_fin, tipo_ciclo in ciclos:
            hora_inicio_obj = datetime.strptime(hora_inicio, "%H:%M")
            hora_fin_obj = datetime.strptime(hora_fin, "%H:%M")

            # Si el ciclo abarca la medianoche (ejemplo de 19:00 a 05:59)
            if hora_inicio_obj > hora_fin_obj:
                # Verificar si la hora está en el rango nocturno cruzando medianoche
                if hora_obj >= hora_inicio_obj or hora_obj <= hora_fin_obj:
                    return tipo_ciclo
            else:
                # Verificar si la hora está dentro del rango diurno/nocturno
                if hora_inicio_obj <= hora_obj <= hora_fin_obj:
                    return tipo_ciclo
        return "Desconocido"  # Si no se encuentra en ningún ciclo
    except Exception as e:
        print(f"Error al determinar el ciclo: {e}")
        return "Desconocido"


# Modificación de la función cargar_excel_y_procesar
def cargar_excel_y_procesar():
    # Seleccionar el archivo Excel
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo de Excel",
        filetypes=(("Archivos de Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*"))
    )
    if archivo:
        try:
            # Leer el archivo Excel
            df = pd.read_excel(archivo)

            # Validar que las columnas requeridas estén presentes
            columnas_necesarias = ['ID Empleado', 'Nombre', 'Apellido', 'Cédula', 'Fecha', 'Hora']
            if not all(col in df.columns for col in columnas_necesarias):
                messagebox.showerror("Error", "El archivo no tiene las columnas necesarias.")
                return

            # Agrupar por 'ID Empleado' y 'Fecha', combinar horas con fecha y limitar a 2 horas
            df['Hora'] = df['Hora'].astype(str)
            resultado = (
                df.groupby(['ID Empleado', 'Fecha'], as_index=False)
                .agg({
                    'Cédula': 'first',
                    'Nombre': 'first',
                    'Apellido': 'first',
                    'Hora': lambda x: ' / '.join(x[:2])  # Limitar a 2 horas
                })
            )

            # Mantener solo la fecha en la columna 'Fecha' y combinar solo las horas en la columna 'Hora'
            resultado['Hora'] = resultado['Hora'].str.strip()
            resultado = resultado[['ID Empleado', 'Cédula', 'Nombre', 'Apellido', 'Fecha', 'Hora']]

            # Agregar la columna de "Tiempo Total (Horas y Minutos)"
            resultado['Tiempo Total (Horas y Minutos)'] = resultado['Hora'].apply(
                lambda x: calcular_tiempo_total(x.split(' / ')[0], x.split(' / ')[-1]))
            resultado['Tiempo Total (Horas y Minutos)'] = resultado['Tiempo Total (Horas y Minutos)'].apply(
                lambda x: f"{int(x)}:{int((x % 1) * 60):02}"  # Convertir a formato HH:MM
            )

            # Agregar la columna de "Tiempo de trabajo limpio" restando 30 minutos
            resultado['Día de la semana'] = resultado['Fecha'].apply(lambda x: datetime.strptime(str(x), "%Y-%m-%d").strftime('%A'))

            # Mapeo de días de la semana al español
            dias_en_espanol = {
                'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
                'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
            }

            resultado['Día de la semana'] = resultado['Día de la semana'].map(dias_en_espanol)

            # Agregar la columna "Tiempo de trabajo limpio"
            resultado['Tiempo de trabajo limpio'] = resultado.apply(
                lambda row: restar_30_minutos(row['Tiempo Total (Horas y Minutos)'], row['Día de la semana']), axis=1
            )

            # Obtener los ciclos de la base de datos
            ciclos_db = obtener_ciclos_db()

            # Agregar la columna "Ciclos"
            resultado['Ciclos'] = resultado['Hora'].apply(lambda x: determinar_ciclo(x.split(' / ')[0], ciclos_db))

            # Agregar la columna "Festivos"
            resultado = agregar_columna_festivos(resultado)


            # Definir el orden de las columnas deseado
            orden_columnas = [
                'ID Empleado', 'Cédula', 'Nombre', 'Apellido', 'Fecha', 'Hora', 'Día de la semana',
                'Ciclos', 'Festivos','Tiempo Total (Horas y Minutos)', 'Tiempo de trabajo limpio'
            ]

            # Reordenar las columnas del DataFrame según el orden deseado
            resultado = resultado[orden_columnas]







            # Guardar el resultado en un nuevo archivo Excel
            archivo_salida = archivo.replace(".xlsx", "_procesado.xlsx")
            resultado.to_excel(archivo_salida, index=False, engine='openpyxl')

            # Formatear el archivo Excel
            formatear_excel(archivo_salida)

            # Mostrar mensaje de éxito
            messagebox.showinfo("Éxito", f"Archivo procesado y guardado como: {archivo_salida}")

            # Abrir el archivo procesado automáticamente
            os.startfile(archivo_salida)
        except Exception as e:
            messagebox.showerror("Error", f"Hubo un problema procesando el archivo: {e}")
    else:
        messagebox.showwarning("Cancelado", "No se seleccionó ningún archivo.")


# Función para formatear el archivo Excel
def formatear_excel(ruta_archivo):
    # Cargar el archivo Excel
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    # Aplicar color amarillo fuerte a la primera fila
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo fuerte
    for col in ws.iter_cols(min_row=1, max_row=1):  # Primera fila
        for cell in col:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar un poco más

    # Centrar el contenido de todas las celdas
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Guardar cambios
    wb.save(ruta_archivo)


# Ventana principal para cargar y procesar el Excel
def ventana_cargar_excel():
    ventana = tk.Tk()
    ventana.title("Cargar y Procesar Excel")
    ventana.configure(bg="#2C2F38")

    # Tamaño y posición de la ventana
    ancho_ventana = 300
    alto_ventana = 150
    screen_width = ventana.winfo_screenwidth()
    screen_height = ventana.winfo_screenheight()
    pos_x = int((screen_width - ancho_ventana) / 2)
    pos_y = int((screen_height - alto_ventana) / 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{pos_x}+{pos_y}")

    # Botón para cargar y procesar el archivo Excel
    boton_cargar = tk.Button(ventana, text="Cargar Excel", command=cargar_excel_y_procesar, font=("Arial", 14),
                             bg="#4CAF50", fg="white", relief="solid", width=20, height=2)
    boton_cargar.pack(pady=40)

    ventana.mainloop()


# Ejecutar la ventana principal
if __name__ == "__main__":
    ventana_cargar_excel()